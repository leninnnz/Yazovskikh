from DataSet_class import DataSet
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
import matplotlib.pyplot as plt
import matplotlib
from jinja2 import Environment, FileSystemLoader
import pdfkit
import openpyxl

border = Border(left=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000'),
                top=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'),)

font = Font(bold=True)


class Report:
    """
    Класс для вывода всей полученной статистики в разных форматах

    Атрибуты
    ------------------------------------------------------------------------
    __border: Border
        Объект для хранения стиля границ ячеек в таблице Excel
    __headline_font: Font
        Объект для хранения тиля шрифта для ечеек в таблице Excel
    """
    def __init__(self, border: Border = border, font: Font = font):
        """
        Инициализирует объект
        :param border:
            Стиль для границ
        :param font:
            Стилья для шрифта
        """
        self.__border = border
        self.__headline_font = font

    def generate_excel(self, data: DataSet, salaries_by_town: {str: float}, rates_by_town: {str: float}, current_vacancy_name: str):
        """
        Генерирует Excel файл со всеми готовыми листами и таблицами составленными из данных
        :param data: DataSet
            Данные для создания таблиц
        :param salaries_by_town: {str: float}
            Дополнительные данные для создания таблиц(отобранные и отсортированные)
            Распределение зарплат по городам (ТОП-10)
        :param rates_by_town:{str: float}
            Дополнительные данные для создания таблиц(отобранные и отсортированные)
            Доли вакансий по городам (ТОП-10)
        :param current_vacancy_name: str
            Название интересующей нас профессии
        :return: void
        """
        wb = Workbook()
        years_sheet = wb.active
        self.__fill_year_sheet(years_sheet, data, current_vacancy_name)

        town_sheet = wb.create_sheet("Статистика по городам")
        self.__fill_town_sheet(town_sheet, salaries_by_town, rates_by_town)

        for column in ["A", "B", "C", "D", "E"]:
            years_sheet.column_dimensions[column].width =\
                max(list(map(lambda cell: len(str(cell.value)), years_sheet[column]))) + 2
            town_sheet.column_dimensions[column].width = \
                max(list(map(lambda cell: len(str(cell.value)), town_sheet[column]))) + 2

        wb.save("report.xlsx")

    def __fill_year_sheet(self, sheet, data, current_vacancy_name):
        """
        Заполняет лист данными о распределении по годам
        :param sheet: Sheet
            Лист, в котором создаются таблицы
        :param data: DataSet
            Данные для заполнения
        :param current_vacancy_name: str
            Название интересующей нас профессии
        :return: void
        """
        sheet.title = "Статистика по годам"
        sheet.append(["Год", "Средняя зарплата", "Средняя зарплата - " + current_vacancy_name,
                      "Количество вакансий", "Количество вакансий - " + current_vacancy_name])

        for year in data.salaries_by_year.keys():
            sheet.append([year, data.salaries_by_year[year], data.current_salaries_by_year[year],
                          data.vacancies_count_by_year[year], data.current_count_by_year[year]])

        for row in sheet:
            for cell in row:
                if cell.row == 1:
                    cell.font = self.__headline_font
                cell.border = self.__border

    def __fill_town_sheet(self, sheet, salaries_by_town: {str: float}, rates_by_town: {str: float}):
        """
        Заполняет лист данными о распределении по городам
        :param sheet: Sheet
            Лист, в котором создаются таблицы
        :param salaries_by_town: {str: float}
            Словарь с данными о распределении зарплат по городам
        :param rates_by_town: {str: float}
            Словарь с данными о распределении долей вакансий по городам
        :return: void
        """
        sheet.append(["Город", "Уровень зарплат", " ", "Город", "Доля вакансий"])

        town_rows = []
        for town_item in salaries_by_town.items():
            town_rows.append([town_item[0], town_item[1], " "])

        for i, town_item in enumerate(rates_by_town.items()):
            town_rows[i] += [town_item[0], town_item[1]]
            sheet.append(town_rows[i])

        for row in sheet:
            for cell in row:
                if cell.column == 3:
                    continue
                if cell.row == 1:
                    cell.font = self.__headline_font
                cell.border = self.__border
                if cell.column == 5:
                    cell.number_format = FORMAT_PERCENTAGE_00

    def generate_image(self, data:DataSet, salaries_by_town: {str: float},
                       rates_by_town: {str: float}, current_vacancy_name: str):
        """
        Генерирует png изображение с графиками построенными на основе данных
        :param data: DataSet
            Данные для создания графиков
        :param salaries_by_town: {str: float}
            Дополнительные данные для создания графиков(отобранные и отсортированные)
            Распределение зарплат по городам (ТОП-10)
        :param rates_by_town:{str: float}
            Дополнительные данные для создания графиков(отобранные и отсортированные)
            Доли вакансий по городам (ТОП-10)
        :param current_vacancy_name: str
            Название интересующей нас профессии
        :return: void
        """
        matplotlib.rc('xtick', labelsize=8)
        matplotlib.rc('xtick', labelsize=8)
        plt.figure(figsize=(100, 100))
        fig, axs = plt.subplots(2, 2)

        self.__generate_salary_by_year_for_town_graph(data, axs[0, 0], current_vacancy_name)
        self.__generate_count_by_year_for_town_graph(data, axs[0, 1], current_vacancy_name)
        self.__generate_salary_by_town_graph(salaries_by_town, axs)
        self.__generate_rates_by_town_graph(rates_by_town, axs)

        fig.tight_layout()
        fig.savefig('graph.png')

    def generate_town_image(self, data: DataSet) -> str:
        """
        Генерирует png изображение с графиками построенными на основе данных
        :param data: DataSet
            Данные для создания графиков
        :return: str
            Название файла, в который был сохранен график
        """
        matplotlib.rc('xtick', labelsize=8)
        matplotlib.rc('xtick', labelsize=8)
        plt.figure(figsize=(100, 100))
        fig, axs = plt.subplots(2,2)

        sorted_salaries_by_town = dict(
            sorted(data.salaries_by_town.items(), key=lambda item: item[1], reverse=True)[0:10])
        sorted_vacancies_by_rate = dict(sorted(data.vacancies_rate_by_town.items(),
                                               key=lambda item: item[1], reverse=True)[0:10])

        self.__generate_salary_by_year_for_town_graph(data, axs[0, 0], data.current)
        self.__generate_count_by_year_for_town_graph(data, axs[0, 1], data.current)
        self.__generate_salary_by_town_graph(sorted_salaries_by_town, axs[1, 0])
        self.__generate_rates_by_town_graph(sorted_vacancies_by_rate, axs[1, 1])

        file_name = 'town_graph.png'
        fig.tight_layout()
        fig.savefig(file_name)
        return file_name

    def __generate_salary_by_year_for_town_graph(self, data: DataSet, ax, current_vacancy_name: str):
        """
        Создает график показывающий динамику зарплат по годам
        :param data: DataSet
            Данные для создания графика
        :param axs: [axis]
            Массив отдельных графиков на одной фигуре
        :param current_vacancy_name: str
            Название интересующей нас профессии
        :return: void
        """
        x = list(map(lambda year: int(year), data.salaries_by_year.keys()))
        ax.bar(list(map(lambda c: c - 0.35, x)), data.salaries_by_year.values(), width=0.35, label="средняя з/п")
        ax.bar(x, data.current_salaries_by_year.values(), width=0.35, label="з/п " + current_vacancy_name)
        ax.set_title(f'Уровень зарплат по годам \n {data.current}({data.region})')
        ax.set_xticks(x, data.salaries_by_year.keys(), rotation=90, fontsize=8)
        ax.legend(fontsize=8)
        ax.grid(axis='y')

    def __generate_count_by_year_for_town_graph(self, data: DataSet, ax, current_vacancy_name: str):
        """
        Создает график показывающий динамику количества вакансий по годам
        :param data: DataSet
            Данные для создания графика
        :param axs: [axis]
            Массив отдельных графиков на одной фигуре
        :param current_vacancy_name: str
            Название интересующей нас профессии
        :return: void
        """
        x = list(map(lambda year: int(year), data.vacancies_count_by_year.keys()))
        ax.bar(list(map(lambda c: c - 0.35, x)), data.vacancies_count_by_year.values(),
                      width=0.35, label="Количество вакансий")
        ax.bar(x, data.current_count_by_year.values(),
                      width=0.35, label="Количество вакансий " + current_vacancy_name)
        ax.set_title(f'Количество вакансий по годам \n {data.current}({data.region})')
        ax.set_xticks(x, data.vacancies_count_by_year.keys(), rotation=90, fontsize=8)
        ax.legend(fontsize=8)
        ax.grid(axis='y')

    def __generate_salary_by_town_graph(self, salaries_by_town: {str: float}, ax):
        """
        Создает график распределения зарплат по городам
        :param salaries_by_town: {str: float}
            Распределение зарплат по городам (ТОП-10)
        :param axs: [axis]
            Массив отдельных графиков на одной фигуре
        :return: void
        """
        x = list(map(lambda town: town.replace(" ", "\n").replace("-", "-\n"), salaries_by_town.keys()))
        ax.barh(x, salaries_by_town.values(), align='center')
        ax.set_title('Уровень зарплат по городам')
        ax.set_yticks(range(10))
        ax.set_yticklabels(x, fontsize=6)
        ax.grid(axis='x')
        ax.invert_yaxis()

    def __generate_rates_by_town_graph(self, rates_by_town: {str: float}, ax):
        """
        Создает график показывающий доли вакансий в разных городах
        :param rates_by_town: {str: float}
            Доли вакансий по городам (ТОП-10)
        :param axs: [axis]
            Массив отдельных графиков на одной фигуре
        :return: void
        """
        values = [1 - sum(rates_by_town.values())] + list(rates_by_town.values())
        labels = ["Другие"] + list(rates_by_town.keys())
        params_tuple = ax.pie(values, labels=labels)
        ax.set_title('Доля вакансий по городам')
        [_.set_fontsize(6) for _ in params_tuple[1]]

    def generate_pdf(self, current_vacancy_name: str, image_file: str, tables_file: str):
        """
        Генерирует отчет в виде PDF файла используя готовые файлы графиков и таблиц
        :param current_vacancy_name: str
            Название интересующей нас профессии
        :param image_file: str
            Название png файла с графиками
        :param tables_file: str
            Название Excel файла с таблицами
        :return: void
        """
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")

        xfile = openpyxl.load_workbook(tables_file)
        years_headlines, years_values, towns_salaries_headlines,\
            towns_rates_headlines, towns_salaries_values, towns_rates_values = [], [], [], [], [], []

        years_table = xfile["Статистика по годам"]
        years_headlines = years_table[1]
        years_values = [row for row in years_table if row != years_table[1]]

        self.__get_towns_table(xfile, towns_salaries_headlines, towns_rates_headlines,
                               towns_salaries_values, towns_rates_values)

        pdf_template = template.render({'vacancy_name': current_vacancy_name, 'image_file': image_file,
                                        'years_headlines': years_headlines, 'years_values': years_values,
                                        'towns_salaries_headlines': towns_salaries_headlines,
                                        'towns_salaries_values': towns_salaries_values,
                                        'towns_rates_headlines': towns_rates_headlines,
                                        'towns_rates_values': towns_rates_values})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        options = {'enable-local-file-access': None}
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options=options)

    def generate_town_pdf(self, data: DataSet, image_file: str):
        """
        Генерирует отчет в виде PDF файла используя готовые файлы графиков
        :param data: DataSet
            Датасет с всей информацией
        :param image_file: str
            Название png файла с графиками
        :return: void
        """
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("town_pdf_template.html")

        years_headlines, years_values, towns_salaries_headlines, \
            towns_rates_headlines = [], [], [], []

        years_headlines = ['Год', f'Средняя зарплата - {data.current}({data.region})',
                           f'Количество вакансий {data.current}({data.region})']

        for year in data.salaries_by_year_for_town.keys():
            years_values.append([year, data.salaries_by_year_for_town[year], data.count_by_year_for_town[year]])

        towns_salaries_headlines = ['Город', 'Уровень зарплат']
        towns_rates_headlines = ['Город', 'Доля вакансий']

        towns_tables = self.__get_towns_table(data)

        pdf_template = template.render({'vacancy_name': data.current, 'image_file': image_file,
                                        'years_headlines': years_headlines, 'years_values': years_values,
                                        'towns_salaries_headlines': towns_salaries_headlines,
                                        'towns_salaries_values': towns_tables[0],
                                        'towns_rates_headlines': towns_rates_headlines,
                                        'towns_rates_values': towns_tables[1]})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        options = {'enable-local-file-access': None}
        pdfkit.from_string(pdf_template, 'town_report.pdf', configuration=config, options=options)

    def __get_towns_table(self, data: DataSet):
        """
        Заполняет таблицу данными для вставки в html шаблон
        :param data: DataSet
            Датасет со всей информацией
        :return: ([[str]], [[str]])
            Возвращает кортеж со списками строк для таблиц
        """
        sorted_salaries_by_town = dict(
            sorted(data.salaries_by_town.items(), key=lambda item: item[1], reverse=True)[0:10])
        sorted_vacancies_by_rate = dict(sorted(data.vacancies_rate_by_town.items(),
                                               key=lambda item: item[1], reverse=True)[0:10])

        towns_salaries_values = [[item[0], item[1]] for item in sorted_salaries_by_town.items()]
        towns_rates_values = [[item[0], item[1]] for item in sorted_vacancies_by_rate.items()]

        return (towns_salaries_values, towns_rates_values)
